""" this project is about taking a xlsx file and iterating over all of Its rows, then sending emails to
client that appear to be missing a paying date on the xlsx file. 
I made this whole thing from scrath without looking at the book, It was hell of work,
but I feel proud after building something functionning all by myself!
"""


import os
original_cwd = os.getcwd()
os.chdir(r"C:\Users\maroc\Downloads\automate_online-materials") #go to the xlsx file location
xlsx_fileName = "duesRecords.xlsx"
########### xlsx stuff
import openpyxl
workbook = openpyxl.load_workbook(xlsx_fileName)
sheet = workbook.active
max_columns = sheet.max_column
max_rows = sheet.max_row
#############
general_list = []  #the big list
sub_list = []  #the small list inside the big list
for ro_w in sheet.iter_rows(min_row=1,max_row=max_rows,min_col=1,max_col=max_columns): #iter in every row
    for sub_row in ro_w:
        sub_list.append(sub_row.value)    # every row = list , kinda of like csv

        
    general_list.append(sub_list)   # every list of a row = a part of a bigger list, looks even more like a csv
    sub_list = []  #empty the small list so I doesnt get more value on top of It
 # and now we make use of that list
client_details_list = []
client_details_sublist = []
for client_details in general_list[1:]:  # the whole list minus the header
    client_details_sublist.append(client_details[0])  # the name of the client
    client_details_sublist.append(client_details[1])  # the email of the client
    for x in range(len(client_details)):   # no need to add -1 to len, why? because the last element of range is not included you brick head
        if client_details[x] is None:   #check If there are unpaid dates
            client_details_sublist.append(general_list[0][x])   #add unpaid dates to the list
    client_details_list.append(client_details_sublist)   #add the sublist to the big list
    client_details_sublist = []  #empty the sublist

    



    

        
            
            


#### basic smtpblip stuff
import smtplib
conn = smtplib.SMTP('smtp.gmail.com', 587)
conn.ehlo()
conn.starttls()
email = input("enter your email: ")
password = input("enter your password: ")
conn.login(email, password)
####
for client_lista in client_details_list:   #iter in the list of clients with their details
    if len(client_lista) > 2:  # this checks if client has unpaid dates
        conn.sendmail(email,client_lista[1],f"Subject: Missing payement \n\n Dear {client_lista[0]}, \n It seems that you are missing a payement for {", ".join(client_lista[2:])}. Please make sure to pay as soon as possible. \n\n Best Regards, \n\n Carlos")

conn.quit()