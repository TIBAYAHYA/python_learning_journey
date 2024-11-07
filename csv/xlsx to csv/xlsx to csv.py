#this program is supposed to convert xlsx files to csv files
import os,sys
import openpyxl
import csv

for xlsx_file in [f for f in os.listdir(".") if f.endswith(".xlsx")]: # filter files that end with .xlsx as extension
    workbook =openpyxl.load_workbook(xlsx_file)
    sheet_names = workbook.sheetnames
#as Im writing this code, every xlsx in cwd has only 1 sheet, but for future uses,
# I wanted to make sure that It iterates over every sheet, and creates a csv file with
#the old xlsx name + sheet name
    for sheet in sheet_names:
        sheet_csv_data = []
        current_sheet = workbook[sheet]
        new_csv_name = xlsx_file[:-5]+"_"+sheet+".csv"
        csv_file_open = open(new_csv_name,"w",newline="")
        csv_writer = csv.writer(csv_file_open)
        for rowNum in range(1, current_sheet.max_row + 1):         # x of row numbers
            row_data = []
            for columnNum in range(1,current_sheet.max_column+1):    # y of column numbers
                row_data.append(current_sheet.cell(row=rowNum,column=columnNum).value)
            sheet_csv_data.append(row_data)
        csv_writer.writerows(sheet_csv_data)
        csv_file_open.close()