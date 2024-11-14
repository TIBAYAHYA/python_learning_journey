

import os
original_cwd = os.getcwd()
os.chdir(r"C:\Users\maroc\Downloads\automate_online-materials")
xlsx_fileName = "duesRecords.xlsx"
import openpyxl
workbook = openpyxl.load_workbook(xlsx_fileName)
sheet = workbook.active
max_columns = sheet.max_column
max_rows = sheet.max_row
general_list = []
sub_list = []
for ro_w in sheet.iter_rows(min_row=1,max_row=max_rows,min_col=1,max_col=max_columns):
    for sub_row in ro_w:
        sub_list.append(sub_row.value)
        
    general_list.append(sub_list)
    sub_list = []

client_list = general_list[1:-1]
for client in client_list:
    if client == None:
        client
